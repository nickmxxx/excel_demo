from openpyxl import load_workbook
wb = load_workbook("example.xlsx")
ws = wb.active 
ws.title = "demo"
print(wb.sheetnames)
ws = wb["demo"]
print(ws.title)
ws1 = wb.create_sheet("demo1")
print(wb.sheetnames)
wb.save("example1.xlsx")