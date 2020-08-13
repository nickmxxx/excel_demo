import openpyxl
wb = openpyxl.Workbook().save("example.xlsx")
wb = openpyxl.load_workbook("example.xlsx")
# print(wb.sheetnames)
ws = wb.active
ws.title = "demo"
ws1 = wb.create_sheet("demo1")
print(wb.sheetnames)
# ws.title  # 'Sheet'
# ws.title = "demo"
# wb.sheetnames  # 'demo'
# ws = wb["demo"]
# ws1 = wb.create_sheet("demo1")
# wb.sheetnames  # ['demo', 'demo1']
# ws['A1'] = 0
# ws.cell(row=1, column=1, value=0)
# print(ws['A1'].value)
for i in range(1, 20):
    for j in range(1, 20):
        ws.cell(i, j, value=i*j)
wb.save("example.xlsx")
