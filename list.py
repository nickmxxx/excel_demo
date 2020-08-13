import openpyxl
import os  # 用来统计excel文件数量
print(os.path.abspath('.'))  # 打开文件夹统计xlsx个数
n = [
    i for i in os.listdir('.')  # 循环查找"."文件
    if os.path.isfile(i) and os.path.splitext(i)[1] == '.xlsx'
]   # 如果是.xlsx计数 加一得到这些文件 放入字典中【】
# print(n)
# print(len(n))
# print(n[0])
wb0 = openpyxl.load_workbook(filename=n[0])  # 打开文件judge1
ws0 = wb0.active
# 第一个评分judge1表作为基准表
print(ws0.cell(3, 1).value)
wb1 = openpyxl.Workbook().save("list.xlsx")
wb1 = openpyxl.load_workbook("list.xlsx")
ws1 = wb1.active
# 两个表格同时使用 将基准表的值赋给最终总表
for i in range(1, ws0.max_row + 1):
    ws1.cell(i, 1).value = ws0.cell(i, 1).value
# 赋值学号

for i in range(1, len(n)):
    print("judge"+str(i))
    ws1.cell(1, i+1).value = "judge"+str(i)
# 赋值judge1，judge2，judge3表头
for i in range(2, ws0.max_row+1):
    ws1.cell(i, 2).value = ws0.cell(i, 2).value

# 此处开始模糊匹配,找唯一确定标识符：此处为学号 用循环从judge2开始
for i in range(1, 3):  # judge2，judge3都循环给总部
    wb2 = openpyxl.load_workbook(n[i])
    ws2 = wb2.active
    for j in range(2, ws1.max_row+1):
        for k in range(2, ws2.max_row+1):
            if(ws2.cell(j, 1).value == ws1.cell(k, 1).value):  # 判断学号相同
                print(ws2.cell(j, 2).value)  # 相同赋值给总表
                ws1.cell(j, i+2).value = ws2.cell(j, 2).value
wb1.save("list.xlsx")
