import openpyxl
import os

# 此程序是为了找出两个表格进行比较来查看谁没有提交文件
print(os.path.abspath('.'))  # 打开文件夹统计xlsx个数
n = [
    i for i in os.listdir('.')  # 循环查找"."文件
    if os.path.isfile(i) and os.path.splitext(i)[1] == '.xlsx'
]  # 如果是.xlsx计数 加一得到这些文件 放入字典中【】
print(n)
print(len(n))
print(n[0])
wb0 = openpyxl.load_workbook(filename=n[0])
wb1 = openpyxl.load_workbook(filename=n[1])
ws0 = wb0.active
ws1 = wb1.active
all_list = []
done_list = []

for i in range(1, 34):
    # print(ws0.cell(i, 1).value)
    all_list.append(ws0.cell(i, 1).value)
# print("\n")
# for i in range(0, 33):
#     print(all_list[i])

for i in range(4, 37):
    # print(ws1.cell(i, 1).value)
    done_list.append(ws1.cell(i, 1).value)
# print("\n")
# for i in range(0, 30):
#     print(done_list[i])
print(len(all_list))
print(len(done_list))
matched = []
for i in range(0, len(all_list)):
    matched.append(0)
# for i in range(0, len(all_list)):
all_student = len(all_list)
print(len(matched))
for i in range(0, len(done_list)):
    for j in range(0, all_student-i):
        if done_list[i] == all_list[j]:
            all_list.remove(all_list[j])
            break
print(all_list)
