import openpyxl
import os  # 用来统计excel文件数量
# import pandas as pd
# from pandas import DataFrame

print(os.path.abspath('.'))  # 打开文件夹统计xlsx个数
n = [
    i for i in os.listdir('.')  # 循环查找"."文件
    if os.path.isfile(i) and os.path.splitext(i)[1] == '.xlsx'
]  # 如果是.xlsx计数 加一得到这些文件 放入字典中【】
print(n)
print(len(n))
print(n[0])

wb_zhiyu = openpyxl.load_workbook(filename=n[3])  # 打开文件智育三年排名
ws_zhiyu = wb_zhiyu.active
# 智育表作为基准表
sheets_zhiyu = wb_zhiyu.sheetnames
print(sheets_zhiyu)
print(ws_zhiyu.cell(1, 1).value)
# wb_zonghe = openpyxl.Workbook().save("zonghe.xlsx")
# wb_zonghe = openpyxl.load_workbook("zonghe.xlsx")
# ws_zonghe = wb_zonghe.active
# 打开大一，大二，大三的德育表
wb_deyu_dayi = openpyxl.load_workbook(filename=n[0])
ws_deyu_dayi = wb_deyu_dayi.active

wb_deyu_daer = openpyxl.load_workbook(filename=n[2])
ws_deyu_daer = wb_deyu_daer.active

wb_deyu_dasan = openpyxl.load_workbook(filename=n[1])
ws_deyu_dasan = wb_deyu_dasan.active

wb_jingsai = openpyxl.load_workbook(filename=n[4])
ws_jingsai = wb_jingsai.active
# 以学号为标志进行查找
for i in range(1, 150):
    for j in range(1, 500):
        print(int(ws_zhiyu.cell(i, 2).value))
        print(int(ws_deyu_dayi.cell(j, 1).value))
        print(int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_dayi.cell(j, 1).value))
        if int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_dayi.cell(j, 1).value):
            ws_zhiyu.cell(i, 6).value = ws_deyu_dayi.cell(j, 4).value
            print(ws_zhiyu.cell(i, 6).value)
            break
    print(ws_zhiyu.cell(i, 6).value)

for i in range(1, 150):
    for j in range(1, 500):
        print(int(ws_zhiyu.cell(i, 2).value))
        print(int(ws_deyu_daer.cell(j, 1).value))
        print(int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_daer.cell(j, 1).value))
        if int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_daer.cell(j, 1).value):
            ws_zhiyu.cell(i, 7).value = ws_deyu_daer.cell(j, 4).value
            print(ws_zhiyu.cell(i, 7).value)
            break
    print(ws_zhiyu.cell(i, 7).value)

for i in range(1, 150):
    for j in range(1, 531):
        print(int(ws_zhiyu.cell(i, 2).value))
        print(int(ws_deyu_dasan.cell(j, 1).value))
        print(int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_dasan.cell(j, 1).value))
        if int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_dasan.cell(j, 1).value):
            ws_zhiyu.cell(i, 8).value = ws_deyu_dasan.cell(j, 4).value
            break
    print(ws_zhiyu.cell(i, 8).value)

for i in range(1, 150):
    for j in range(1, 110):
        print(int(ws_zhiyu.cell(i, 2).value))
        print(int(ws_jingsai.cell(j, 2).value))
        print(int(ws_zhiyu.cell(i, 2).value) == int(ws_deyu_dasan.cell(j, 1).value))
        if int(ws_zhiyu.cell(i, 2).value) == int(ws_jingsai.cell(j, 2).value):
            ws_zhiyu.cell(i, 9).value = ws_jingsai.cell(j, 4).value
            break
        print(ws_zhiyu.cell(i, 9).value)

wb_zhiyu.save('智育.xlsx')
# 两个表格同时使用 将基准表的值赋给最终总表
# for i in range(1, 47):  # 11为列数，共统计11列
#     ws1.cell(1, i).value = ws0.cell(1, i).value
#     ws1.cell(2, i).value = ws0.cell(2, i).value
# 复制第一，二行赋值
# wb1.save("list.xlsx")
# for i in range(1, len(n)):
#     print("judge"+str(i))
#     ws1.cell(1, i+1).value = "judge"+str(i)
# 赋值judge1，judge2，judge3表头
# for i in range(2, ws0.max_row+1):
#     ws1.cell(i, 2).value = ws0.cell(i, 2).value

# 此处将原来的表值赋给新表

# 规范表格内容，如有相同的内容则直接赋值
# for i in range(3, len(n) + 3):
#     ws1.cell(i, 1).value = "信息与通信工程学院"
#     ws1.cell(i, 5).value = "贺美琛"
# wb1.save("list.xlsx")
# 进行改进，添加排序功能，此处按照学号进行排序
# cel = pd.read_excel("list.xlsx")
# cel = pd.read_excel("list.xlsx", header=1)
# cel.sort_values(by='学号', inplace=True, ascending=True)
# DataFrame(cel).to_excel('list.xlsx')
