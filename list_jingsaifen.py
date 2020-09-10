import openpyxl
import os  # 用来统计excel文件数量
# import pandas as pd
# from pandas import DataFrame

print(os.path.abspath('.'))  # 打开文件夹统计xlsx个数
n = [
    i for i in os.listdir('.')  # 循环查找"."文件
    if os.path.isfile(i) and os.path.splitext(i)[1] == '.xlsx'
]  # 如果是.xlsx计数 加一得到这些文件 放入字典中【】
print(n)
print(len(n))
print(n[0])
wb0 = openpyxl.load_workbook(filename=n[0])  # 打开文件judge1
ws0 = wb0.active
# 第一个评分judge1表作为基准表
print(ws0.cell(1, 1).value)
wb1 = openpyxl.Workbook().save("list.xlsx")
wb1 = openpyxl.load_workbook("list.xlsx")
ws1 = wb1.active
# 两个表格同时使用 将基准表的值赋给最终总表
# for i in range(1, 47):  # 11为列数，共统计11列
#     ws1.cell(1, i).value = ws0.cell(1, i).value
#     ws1.cell(2, i).value = ws0.cell(2, i).value
# 复制第一，二行赋值
# wb1.save("list.xlsx")
# for i in range(1, len(n)):
#     print("judge"+str(i))
#     ws1.cell(1, i+1).value = "judge"+str(i)
# 赋值judge1，judge2，judge3表头
# for i in range(2, ws0.max_row+1):
#     ws1.cell(i, 2).value = ws0.cell(i, 2).value

# 此处将原来的表值赋给新表
for i in range(0, len(n)):  # judge2，judge3都循环给总部
    wb2 = openpyxl.load_workbook(n[i])
    ws2 = wb2.active
    print("这是" + n[i])
    # for j in range(2, ws1.max_row+1):
    #     for k in range(2, ws2.max_row+1):
    #         if(ws2.cell(j, 1).value == ws1.cell(k, 1).value):  # 判断学号相同
    for j in range(1, 47):
        print(ws2.cell(9, j).value)  # 每个分表第三行赋值给总表
        ws1.cell(i + 1, j).value = ws2.cell(9, j).value
wb1.save("list.xlsx")
# 规范表格内容，如有相同的内容则直接赋值
# for i in range(3, len(n) + 3):
#     ws1.cell(i, 1).value = "信息与通信工程学院"
#     ws1.cell(i, 5).value = "贺美琛"
# wb1.save("list.xlsx")
# 进行改进，添加排序功能，此处按照学号进行排序
# cel = pd.read_excel("list.xlsx")
# cel = pd.read_excel("list.xlsx", header=1)
# cel.sort_values(by='学号', inplace=True, ascending=True)
# DataFrame(cel).to_excel('list.xlsx')
